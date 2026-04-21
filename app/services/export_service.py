"""
Export Service / 导出服务

提供证书 Excel 导出功能
"""
from io import BytesIO
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app import db
from app.models import Certificate, ExportTemplate


# 可用导出字段定义
AVAILABLE_FIELDS = {
    'title': '标题',
    'owner_name': '持有人',
    'department': '部门',
    'cert_type': '证书类型',
    'issue_date': '获奖日期',
    'issuer': '颁发机构',
    'created_at': '创建时间',
    'expire_date': '过期日期',
    'description': '描述/备注',
}

# 默认导出的字段
DEFAULT_FIELDS = ['title', 'owner_name', 'department', 'cert_type', 'issue_date', 'issuer', 'created_at']


def get_available_fields() -> List[dict]:
    """获取所有可用导出字段"""
    return [
        {'key': key, 'label': label}
        for key, label in AVAILABLE_FIELDS.items()
    ]


def get_default_template(user_id: int, is_admin: bool = False) -> ExportTemplate:
    """获取用户默认模板

    Args:
        user_id: 用户ID
        is_admin: 是否为管理员（学校管理员），管理员用自己的默认模板
    """
    # 管理员用自己的默认模板
    if is_admin:
        template = ExportTemplate.get_default_for_user(user_id)
        if not template:
            template = ExportTemplate.init_default_template(user_id)
        return template

    # 非管理员使用全局默认模板（管理员设置的）
    global_default = ExportTemplate.get_global_default_template()
    if global_default:
        return global_default

    # 如果没有全局默认，则用用户自己的（兼容旧逻辑）
    template = ExportTemplate.get_default_for_user(user_id)
    if not template:
        template = ExportTemplate.init_default_template(user_id)
    return template


def get_user_templates(user_id: int) -> List[ExportTemplate]:
    """获取用户的所有模板"""
    return ExportTemplate.get_all_for_user(user_id)


def generate_certificate_excel(certificates: List[Certificate], fields: List[str]) -> BytesIO:
    """
    根据指定字段生成证书 Excel 文件

    Args:
        certificates: 证书查询结果列表
        fields: 要导出的字段列表

    Returns:
        BytesIO 对象，可直接用于文件下载
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '证书清单'

    # 设置表头样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    headers = [AVAILABLE_FIELDS.get(f, f) for f in fields]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置列宽
    column_widths = {
        'title': 30,
        'owner_name': 15,
        'department': 15,
        'cert_type': 15,
        'issue_date': 15,
        'issuer': 25,
        'created_at': 20,
        'expire_date': 15,
        'description': 40,
    }
    for col, field in enumerate(fields, start=1):
        width = column_widths.get(field, 15)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # 写入数据行
    for row_idx, cert in enumerate(certificates, start=2):
        for col_idx, field in enumerate(fields, start=1):
            value = _get_field_value(cert, field)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # 生成文件名
    filename = f'证书清单_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    # 保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _get_field_value(cert: Certificate, field: str):
    """获取证书指定字段的值"""
    if field == 'title':
        return cert.title
    elif field == 'owner_name':
        return cert.owner.name if cert.owner else ''
    elif field == 'department':
        return cert.owner.department.name if cert.owner and cert.owner.department else ''
    elif field == 'cert_type':
        return cert.certificate_type.name if cert.certificate_type else ''
    elif field == 'issue_date':
        # issue_date stored in fields JSON
        issue_date = cert.fields.get('issue_date') if cert.fields else None
        if issue_date:
            # Handle both string and dict format
            if isinstance(issue_date, dict):
                issue_date = issue_date.get('value', '')
            return issue_date
        return ''
    elif field == 'issuer':
        # issuer stored in fields JSON
        issuer = cert.fields.get('issuer') if cert.fields else None
        if isinstance(issuer, dict):
            issuer = issuer.get('value', '')
        return issuer or ''
    elif field == 'created_at':
        return cert.created_at.strftime('%Y-%m-%d %H:%M:%S') if cert.created_at else ''
    elif field == 'expire_date':
        # expire_date stored in fields JSON
        expire_date = cert.fields.get('expire_date') if cert.fields else None
        if expire_date:
            if isinstance(expire_date, dict):
                expire_date = expire_date.get('value', '')
            return expire_date
        return ''
    elif field == 'description':
        # description stored in fields JSON
        desc = cert.fields.get('description') if cert.fields else None
        if isinstance(desc, dict):
            desc = desc.get('value', '')
        return desc or ''
    return ''
