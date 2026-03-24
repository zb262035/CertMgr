"""Certificate business logic services."""
import os
import uuid
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from app.extensions import db
from app.models.certificate import Certificate, CertificateType
from app.services.file_storage_service import FileStorageService


class CertificateService:
    """Service for certificate CRUD operations."""

    ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif'}

    @staticmethod
    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in CertificateService.ALLOWED_EXTENSIONS

    @staticmethod
    def create_certificate(user_id, title, certificate_type_id, file, dynamic_fields):
        """Create a new certificate with file upload."""
        if not file or not CertificateService.allowed_file(file.filename):
            raise ValueError('Invalid file type')

        original_filename = secure_filename(file.filename)

        # Save file using FileStorageService
        file_result = FileStorageService.save_file(file)
        file_path = file_result['path']
        file_size = file_result['size']
        file_mime_type = file.content_type

        certificate = Certificate(
            user_id=user_id,
            title=title,
            certificate_type_id=certificate_type_id,
            file_path=file_path,
            original_filename=original_filename,
            file_size=file_size,
            file_mime_type=file_mime_type,
            fields=dynamic_fields
        )
        db.session.add(certificate)
        db.session.commit()
        return certificate

    @staticmethod
    def update_certificate(certificate, title, certificate_type_id, file=None, dynamic_fields=None):
        """Update existing certificate."""
        certificate.title = title
        certificate.certificate_type_id = certificate_type_id

        if file and CertificateService.allowed_file(file.filename):
            FileStorageService.delete_file(certificate.file_path)
            file_result = FileStorageService.save_file(file)
            certificate.original_filename = secure_filename(file.filename)
            certificate.file_path = file_result['path']
            certificate.file_size = file_result['size']
            certificate.file_mime_type = file.content_type

        if dynamic_fields is not None:
            certificate.fields = dynamic_fields

        certificate.updated_at = datetime.utcnow()
        db.session.commit()
        return certificate

    @staticmethod
    def delete_certificate(certificate):
        """Delete certificate and its file."""
        FileStorageService.delete_file(certificate.file_path)
        db.session.delete(certificate)
        db.session.commit()


class ExcelImportService:
    """Service for auto-parsing Excel files for batch certificate import.

    No template required - system automatically detects column headers.
    """

    # Column name mappings (Excel header -> Certificate field)
    COLUMN_MAPPINGS = {
        'title': ['标题', '证书名称', '名称', 'name', 'title'],
        'type': ['类型', '证书类型', 'type', 'certificate_type'],
        'competition_name': ['比赛名称', '比赛', 'competition'],
        'award_level': ['奖项', '获奖等级', 'level', 'award_level', '奖级'],
        'award_date': ['获奖日期', '获奖时间', 'award_date', '获奖日期'],
        'organizer': ['主办单位', '组织单位', 'organizer', '主办方'],
        'certificate_number': ['证书编号', '编号', 'number', 'certificate_number', '证书号'],
        'honor_title': ['荣誉名称', '荣誉', 'honor', '荣誉名称'],
        'grant_date': ['授予日期', '授予时间', 'grant_date'],
        'grantor': ['授予单位', 'grantor', '授予机构'],
        'reason': ['获得原因', '原因', 'reason'],
        'certificate_name': ['证书名称', '证书名', 'certificate_name'],
        'issue_date': ['发证日期', '颁发日期', 'issue_date'],
        'expiry_date': ['有效期', '有效期至', 'expiry_date', '过期日期'],
        'issuing_authority': ['发证机构', '颁发机构', 'issuer', 'issuing_authority', '发证单位'],
        'skill_name': ['职业技能', 'skill', 'skill_name', '职业名称'],
        'skill_level': ['等级', 'skill_level', '级别', '职业技能等级'],
    }

    @classmethod
    def parse_excel(cls, file_stream) -> tuple:
        """Parse Excel file and return list of certificate records.

        Returns:
            Tuple of (records, errors)
        """
        wb = load_workbook(file_stream, read_only=True, data_only=True)
        ws = wb.active

        # Read headers from first row
        headers = [cell.value for cell in ws[1]]
        if not headers or not any(headers):
            wb.close()
            return [], ['Excel 文件为空或第一行没有列标题 / Excel file is empty or first row has no column headers']

        # Auto-detect column mapping
        column_map = cls._detect_columns(headers)

        if not column_map.get('title'):
            wb.close()
            return [], ['未找到标题列（标题、名称、证书名称 等）/ Title column not found (title, name, certificate name, etc.)']

        # Parse data rows
        records = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue  # Skip empty rows

            try:
                record = cls._parse_row(row, headers, column_map)
                if record:
                    records.append(record)
            except Exception as e:
                errors.append(f'第 {row_idx} 行解析错误 / Row {row_idx} parse error: {str(e)}')

        wb.close()
        return records, errors

    @classmethod
    def _detect_columns(cls, headers: list) -> dict:
        """Auto-detect which columns map to which certificate fields."""
        column_map = {}

        for field, aliases in cls.COLUMN_MAPPINGS.items():
            for idx, header in enumerate(headers):
                if header and str(header).strip().lower() in [a.lower() for a in aliases]:
                    column_map[field] = idx
                    break

        return column_map

    @classmethod
    def _parse_row(cls, row: tuple, headers: list, column_map: dict) -> dict:
        """Parse a single row into a certificate record."""
        # Get title
        title_idx = column_map.get('title')
        if title_idx is None or not row[title_idx]:
            return None

        title = str(row[title_idx]).strip()

        # Determine certificate type
        type_idx = column_map.get('type')
        cert_type = None
        if type_idx and row[type_idx]:
            type_name = str(row[type_idx]).strip()
            cert_type = CertificateType.query.filter(
                CertificateType.name.like(f'%{type_name}%')
            ).first()

        # If type not specified, default to first type
        if not cert_type:
            cert_type = CertificateType.query.first()

        # Build dynamic fields from mapped columns
        fields = {}
        for field_name, idx in column_map.items():
            if field_name not in ('title', 'type') and idx < len(row):
                value = row[idx]
                if value is not None:
                    # Handle date objects from Excel
                    if hasattr(value, 'strftime'):
                        value = value.strftime('%Y-%m-%d')
                    elif isinstance(value, (int, float)):
                        value = str(value)
                    else:
                        value = str(value).strip()
                    fields[field_name] = value

        return {
            'title': title,
            'certificate_type_id': cert_type.id if cert_type else None,
            'fields': fields,
        }

    @classmethod
    def import_batch(cls, records: list, user_id: int, overwrite: bool = False) -> tuple:
        """Import batch of certificate records.

        Args:
            records: List of parsed certificate records
            user_id: User ID to associate certificates with
            overwrite: If True, update existing certificates by title+user

        Returns:
            Tuple of (success_count, errors)
        """
        success_count = 0
        errors = []

        for record in records:
            try:
                if overwrite:
                    existing = Certificate.query.filter_by(
                        user_id=user_id,
                        title=record['title']
                    ).first()
                    if existing:
                        existing.fields = record['fields']
                        existing.certificate_type_id = record['certificate_type_id']
                        db.session.commit()
                        success_count += 1
                        continue

                certificate = Certificate(
                    user_id=user_id,
                    title=record['title'],
                    certificate_type_id=record['certificate_type_id'],
                    file_path='',  # No file for imported certificates
                    fields=record['fields']
                )
                db.session.add(certificate)
                db.session.commit()
                success_count += 1
            except Exception as e:
                errors.append(f"导入 '{record.get('title', 'Unknown')}' 失败 / Failed to import '{record.get('title', 'Unknown')}': {str(e)}")

        return success_count, errors

